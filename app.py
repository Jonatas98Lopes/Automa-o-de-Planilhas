import openpyxl
from interface import *

def tem_heading_sheet(sheet):
    return sheet['A1'].value is not None 

workbook = openpyxl.Workbook()
del workbook['Sheet']

new_sheets = []
new_columns = []
new_data = []



adiciona_sheets_planilha_, escolhe_sheet_ = adiciona_sheets_planilha(), None
adiciona_colunas_, escolher_adicionar_dados_ = None, None
salvar_arquivo_, escolher_pagina_adicionar_dados_ = None, None
insere_dados_sheet_ = None

while True:
    window, event, values = sg.read_all_windows()

    if event == sg.WINDOW_CLOSED: break

    if window == adiciona_sheets_planilha_:

        if event == 'Adicionar':
            sheet = values['sheet']
            if not sheet.isspace() and sheet != '':
                window['Continuar'].update(disabled=False)
                print(values['sheet'])
                new_sheets.append(values['sheet'])
                window['sheet'].update('')

        elif event == 'Limpar':
            window['sheet'].update('')

        elif event == 'Continuar':
            for new_sheet in new_sheets:
                workbook.create_sheet(new_sheet)
            adiciona_sheets_planilha_.close()
            escolhe_sheet_ = escolhe_sheet(workbook.sheetnames)
            

    elif window == escolhe_sheet_:
        for key, value in values.items():
            if value:
                name_current_sheet = key
                current_sheet = workbook[key]
        escolhe_sheet_.close()
        adiciona_colunas_ = adiciona_colunas(name_current_sheet)

    elif window == adiciona_colunas_:

        if event == 'Adicionar':
            column = values['column_name']
            if not column.isspace() and column != '':
                window['Continuar'].update(disabled=False)
                print(values['column_name'])
                new_columns.append(values['column_name'])
                window['column_name'].update('')

        elif event == 'Limpar':
                window['column_name'].update('')
        
        elif event == 'Continuar':
            current_sheet.append(new_columns)
            adiciona_colunas_.close()
            escolher_adicionar_dados_ = escolher_adicionar_dados()

    elif window == escolher_adicionar_dados_:
        escolher_adicionar_dados_.close()
        if event == 'Continuar':
            if values['sim']:
                escolher_pagina_adicionar_dados_ = escolher_pagina_adicionar_dados(
                    workbook.sheetnames)
            else: 
                salvar_arquivo_ = salvar_arquivo()
                
    elif window == escolher_pagina_adicionar_dados_:
        if event == 'Continuar':
            escolher_pagina_adicionar_dados_.close()
            for key, value in values.items():
                if value:
                    current_sheet = workbook[key]
                    break
            insere_dados_sheet_ = insere_dados_sheet()
                    
    elif window == insere_dados_sheet_:

        if event == 'Adicionar dado':
            dado = values['data'] 
            if not dado.isspace() and dado != '':
                new_data.append(dado)
                print(f'{dado},', end=' ')
                window['data'].update('')
                window['Continuar'].update(disabled=False)

        elif event == 'Limpar': 
            window['data'].update('')

        elif event == 'Enviar linha': 
            if len(new_data) > 0:
                current_sheet.append(new_data)
                new_data.clear()
                print()
                print(30 * '=')
                print('LINHA ADICIONADA')
                print(30 * '=')


        elif event == 'Continuar': 
            insere_dados_sheet_.close()
            salvar_arquivo_ = salvar_arquivo()

    elif window == salvar_arquivo_:
        if event == 'Salvar':
            window['Salvar'].update(disabled=True)
            if values['file_name'].isdigit():
                window['file_name'].update('')
                window['warning_file_name']\
                    .update('⚠VOCÊ DIGITOU APENAS NÚMEROS. NOMES DE ARQUIVOS PRECISAM CONTER LETRAS⚠')
            else:
                file_name = values['file_name'] + values['extension']
                workbook.save(file_name)
                window['Finalizar'].update(disabled=False)
        elif event == 'Finalizar':
            break



""" sheet_choice = input("Digite o nome da página a ser manipulada: ")
current_sheet = workbook[sheet_choice]

new_fields = []
while True:
    new_field = input('Digite um nome de uma coluna para o cabeçalho: ')
    new_fields.append(new_field)


    add_new_field = input('Adicionar mais uma coluna?(s/n): ')
    if add_new_field == 'n': break

current_sheet.append(new_fields)

add_data_in_sheet = input("Adicionar dados a essa planilha?(s/n): ")
if add_data_in_sheet == 's':

    available_sheets = workbook.sheetnames
    print(f'As páginas disponíveis são: {available_sheets}')

    sheet_choice = input("Em qual página devemos adicionar dados? ")
    current_sheet = workbook[sheet_choice]

    while True:
        data = input("Digite os dados a serem adicionados a uma nova linha, separadas por vírgula: ").strip()
        current_sheet.append(data.split(','))

        add_new_data_line = input("Adicionar uma nova linha?(s/n): ")
        if add_new_data_line == 'n': break
    
workbook_name = input("Digite o nome da planilha a ser salva: ").strip()
workbook.save(f'{workbook_name}.xlsx')
print("Planilha criada com sucesso.")
 """